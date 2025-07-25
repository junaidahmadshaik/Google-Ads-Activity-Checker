import streamlit as st
import pandas as pd
import re
import time
from openpyxl import load_workbook

st.set_page_config(page_title="Ad Group Structure & Status Analysis Tool", layout="centered")

# Custom CSS to make file upload button inline with text
st.markdown("""
<style>
    /* Hide the drag and drop area and instructions */
    .stFileUploader [data-testid="stFileUploaderDropzoneInstructions"] {
        display: none !important;
    }
    
    /* Make the entire dropzone container inline and remove styling */
    .stFileUploader [data-testid="stFileUploaderDropzone"] {
        border: none !important;
        background: transparent !important;
        padding: 0 !important;
        margin: 0 !important;
        min-height: auto !important;
        display: inline-block !important;
        vertical-align: middle !important;
    }
    
    /* Style the browse button */
    .stFileUploader [data-testid="stFileUploaderDropzone"] button {
        display: inline-flex !important;
        -webkit-box-align: center !important;
        align-items: center !important;
        -webkit-box-pack: center !important;
        justify-content: center !important;
        font-weight: 400 !important;
        padding: 0.25rem 0.75rem !important;
        border-radius: 0.5rem !important;
        min-height: 2.5rem !important;
        margin: 0px !important;
        line-height: 1.6 !important;
        text-transform: none !important;
        font-size: 16px !important;
        font-family: Source Sans, sans-serif !important;
        color: rgb(250, 250, 250) !important;
        width: auto !important;
        cursor: pointer !important;
        user-select: none !important;
        background-color: rgb(19, 23, 32) !important;
        border: 1px solid rgba(250, 250, 250, 0.2) !important;
    }
    
    .stFileUploader [data-testid="stFileUploaderDropzone"] button:hover {
        background-color: #333 !important;
        border-color: red !important;
        color: red !important;
    }
    
    /* Make the file uploader container inline */
    .stFileUploader {
        display: inline-block !important;
        vertical-align: middle !important;
        width: auto !important;
    }
    
    /* Hide the widget label */
    .stFileUploader [data-testid="stWidgetLabel"] {
        display: none !important;
    }
    
    /* Style the markdown headers to be inline */
    .inline-upload {
        display: flex !important;
        align-items: center !important;
        gap: 0px !important;
    }
    
    .inline-upload h3 {
        margin: 0 !important;
        display: inline-block !important;
    }
</style>
""", unsafe_allow_html=True)

st.title("Google Ads Activity Checker")
st.header("Upload Files")

# Create inline upload sections
col1, col2 = st.columns([3, 1])

with col1:
    st.markdown("### 📘 Upload your `accounts_list.xlsx`")
with col2:
    accounts_file = st.file_uploader("accounts_list", type="xlsx", key="accounts", label_visibility="collapsed")

col1, col2 = st.columns([3, 1])
with col1:
    st.markdown("### 📄 Upload your `keyword_report.xlsx`")
with col2:
    keyword_file = st.file_uploader("keyword_report", type="xlsx", key="keyword", label_visibility="collapsed")

col1, col2 = st.columns([3, 1])
with col1:
    st.markdown("### 📄 Upload your `ad_report.xlsx`")
with col2:
    adgroup_file = st.file_uploader("IM_VDP_ad_group_report", type="xlsx", key="adgroup", label_visibility="collapsed")

# def read_visible_rows_only(file_path):
#     """
#     Read only visible rows from Excel file, filtering out hidden rows
#     """
#     try:
#         # First, try to use openpyxl to identify hidden rows
#         wb = load_workbook(file_path, data_only=True)
#         ws = wb.active

#         if ws is None:
#             raise ValueError("No active worksheet found in the Excel file.")
        
#         # Get visible row indices (1-indexed in openpyxl)
#         visible_rows = []
#         for row_idx in range(2, ws.max_row + 1):
#             if ws.row_dimensions[row_idx].hidden is False:
#                 visible_rows.append(row_idx)  # Convert to 0-indexed for pandas
        
#         # Read all data first
#         df = pd.read_excel(file_path)
       
#         # Filter to only visible rows
#         if visible_rows:
#             # Make sure we don't go beyond the actual dataframe length
#             valid_indices = [i-2 for i in visible_rows if i < len(df)]
#             df = df.iloc[valid_indices]
            
        
#         return df
        
#     except Exception as e:
#         st.warning(f"Could not determine hidden rows using openpyxl: {str(e)}")
#         st.info("Falling back to reading all rows. Please ensure your accounts_list.xlsx only contains the rows you want to process.")
#         # Fallback to regular pandas reading
#         return pd.read_excel(file_path)

def extract_ad_group_pattern(ad_group_text):
    """
    Extract ad group pattern from the full ad group text
    """
    if pd.isna(ad_group_text) or ad_group_text == "":
        return None
    
    # Define the ad group patterns
    patterns = [
        r'New - Lease- \d{4}',
        r'Lease or Other \d{4}',
        r'Other Deal \d{4}',
        r'Finance Other \d{4}',
        r'New - Rebate Deal- \d{4}',
        r'New - Deal - \d{4}',
        r'Other or Finance \d{4}',
        r'Finance or Other \d{4}',
        r'Lease or Finance \d{4}'
    ]
    
    ad_group_text = str(ad_group_text)
    
    for pattern in patterns:
        match = re.search(pattern, ad_group_text)
        if match:
            return ad_group_text
    
    return None

def check_ads_active(row):
    """
    Check if ads are active by looking at headline and description columns
    """
    # Get all headline columns (Headline 1 to Headline 15)
    headline_cols = [col for col in row.index if 'headline' in col.lower()]
    
    # Get all description columns (Description 1 to Description 4)
    description_cols = [col for col in row.index if 'description' in col.lower()]
    
    # Check if any headline or description has content
    for col in headline_cols + description_cols:
        if pd.notna(row[col]) and str(row[col]).strip() != "" and str(row[col]).strip() != "--":
            return True
    
    return False

def process_google_ads_data(accounts_df, keyword_df, adgroup_df):
    """
    Process Google Ads data according to the project requirements
    """
    results = []
    
    # Step 1: Get unique Customer IDs from accounts_list
    if 'Customer ID' not in accounts_df.columns:
        st.error("'Customer ID' column not found in accounts_list.xlsx")
        return []
    
    unique_customer_ids = accounts_df['Customer ID'].dropna().unique()
    
    # Process each unique customer ID
    for customer_id in unique_customer_ids:
        # Get account name for this customer ID
        account_name = accounts_df[accounts_df['Customer ID'] == customer_id]['Account name'].iloc[0] if 'Account name' in accounts_df.columns else ""
        
        # Check if customer ID exists in IM_VDP_ad_group_report
        if 'Customer ID' not in adgroup_df.columns:
            # Customer ID not present in adgroup report - ads and keywords inactive
            results.append({
                'Account name': account_name,
                'Customer ID': customer_id,
                'Campaign': 'No active campaigns',
                'Ad group': 'No IM_VDP ad groups found',
                'ads': 'not active',
                'keywords': 'not active'
            })
            continue
        
        # Get ALL ad groups for this customer ID
        customer_adgroups = adgroup_df[adgroup_df['Customer ID'] == customer_id]
        
        if customer_adgroups.empty:
            # Customer ID not found in adgroup report
            results.append({
                'Account name': account_name,
                'Customer ID': customer_id,
                'Campaign': 'No active campaigns',
                'Ad group': 'No IM_VDP ad groups found',
                'ads': 'not active',
                'keywords': 'not active'
            })
            continue
        
        # Process EACH ad group for this customer ID
        valid_adgroups_found = False
        
        for _, row in customer_adgroups.iterrows():
            ad_group_text = row.get('Ad group', '')
            ad_state = row.get('Ad state', '')
            campaign = row.get('Campaign', '')
            ad_group_id = row.get('Ad group ID', '')


            # Treat "--" as empty/null values
            if str(ad_state).strip() == "--":
                ad_state = ""
            if str(campaign).strip() == "--":
                campaign = ""
            if str(ad_group_id).strip() == "--":
                ad_group_id = ""
            
            # Extract ad group pattern
            ad_group_pattern = extract_ad_group_pattern(ad_group_text)
            
            if ad_group_pattern is None:
                # Ad group doesn't match required structure - skip this ad group
                continue
            
            if ad_state != 'Enabled':
                # Ad group is not enabled - skip this ad group
                continue
            
            # If we reach here, we found a valid ad group
            valid_adgroups_found = True
            
            # Check if ads are active for this specific ad group
            ads_status = 'active' if check_ads_active(row) else 'not active'
            
            # Check if keywords are active for this specific ad group
            keywords_status = 'not active'
            if pd.notna(ad_group_id) and 'Ad group ID' in keyword_df.columns:
                if ad_group_id in keyword_df['Ad group ID'].values:
                    keywords_status = 'active'
            
            # Add a separate result row for each valid ad group
            results.append({
                'Account name': account_name,
                'Customer ID': customer_id,
                'Campaign': campaign if campaign != "" else 'No active campaigns',
                'Ad group': ad_group_pattern,
                'ads': ads_status,
                'keywords': keywords_status
            })
        campaign = customer_adgroups.iloc[0]['Campaign']
        # If no valid ad groups found for this customer ID
        if not valid_adgroups_found:
            results.append({
                'Account name': account_name,
                'Customer ID': customer_id,
                'Campaign': campaign if campaign != "" else 'No active campaigns',
                'Ad group': 'No Ad groups with valid Structure is found',
                'ads': 'not active',
                'keywords': 'not active'
            })
    
    return results


# Initialize session state
if 'results_processed' not in st.session_state:
    st.session_state.results_processed = False
if 'results_data' not in st.session_state:
    st.session_state.results_data = None
if 'accounts_data' not in st.session_state:
    st.session_state.accounts_data = None
if 'keyword_data' not in st.session_state:
    st.session_state.keyword_data = None
if 'adgroup_data' not in st.session_state:
    st.session_state.adgroup_data = None
if 'processing_time' not in st.session_state:
    st.session_state.processing_time = 0

if st.button("Submit"):
    if accounts_file and keyword_file and adgroup_file:
        try:
            # Read Excel files with progress indication
            with st.spinner('Reading Excel files...'):
                # Read accounts file with only visible rows
                accounts_df = pd.read_excel(accounts_file, skiprows=2)
                
                # Read keyword file with proper encoding handling
                keyword_df = pd.read_excel(keyword_file, skiprows=2)
                
                # Read adgroup file
                adgroup_df = pd.read_excel(adgroup_file, skiprows=2)
                
            
            # Display file information
            st.write("### File Information:")
            st.write(f"**Accounts file:** {accounts_df.shape[0]} rows, {accounts_df.shape[1]} columns")
            st.write(f"**Keywords file:** {keyword_df.shape[0]} rows, {keyword_df.shape[1]} columns")
            st.write(f"**Ad Group file:** {adgroup_df.shape[0]} rows, {adgroup_df.shape[1]} columns")
            
            # Process the data with timing
            start_time = time.time()
            results = process_google_ads_data(accounts_df, keyword_df, adgroup_df)
            end_time = time.time()
            
            processing_time = round(end_time - start_time, 2)
            
            # Store results in session state
            st.session_state.results_processed = True
            st.session_state.results_data = results
            st.session_state.accounts_data = accounts_df
            st.session_state.keyword_data = keyword_df
            st.session_state.adgroup_data = adgroup_df
            st.session_state.processing_time = processing_time
            
        except Exception as e:
            st.error(f"Error processing files: {str(e)}")
            st.error("Please check if your files have the required columns and proper formatting:")
            st.write("**accounts_list.xlsx:** 'Client Name'")
            st.write("**keyword_report:** 'Ad group ID' (starting from row 3) - supports .xlsx and .csv")
            st.write("**IM_VDP_ad_group_report.xlsx:** 'Account name', 'Customer ID', 'Campaign', 'Ad group', 'Ad group ID', 'Ad state', 'Headline 1-15', 'Description 1-4' (starting from row 3)")
    else:
        st.warning("Please upload all three files to continue.")

# Display results if they exist in session state
if st.session_state.results_processed and st.session_state.results_data:
    results = st.session_state.results_data
    keyword_df = st.session_state.keyword_data
    adgroup_df = st.session_state.adgroup_data
    processing_time = st.session_state.processing_time
    
    # Display results
    if results:
        st.success(f"✅ Analysis complete in {processing_time} seconds! Found {len(results)} records:")
        result_df = pd.DataFrame(results)
        st.dataframe(result_df, use_container_width=True)

        # Show overall summary statistics
        st.write("### Overall Summary:")
        total_records = len(result_df)
        active_ads = len(result_df[result_df['ads'] == 'active'])
        active_keywords = len(result_df[result_df['keywords'] == 'active'])
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Records", total_records)
        with col2:
            st.metric("Active Ads", f"{active_ads}/{total_records}")
        with col3:
            st.metric("Active Keywords", f"{active_keywords}/{total_records}")
        with col4:
            st.metric("Processing Time", f"{processing_time}s")
        
        # Option to download results
        csv = result_df.to_csv(index=False)
        st.download_button(
            label="Download Results as CSV",
            data=csv,
            file_name="google_ads_activity_report.csv",
            mime="text/csv"
        )
        

        st.write("---")
        # Add dropdown for account analysis
        st.write("### Account Analysis:")
        
        # Get unique account names from results (Client Name column)
        valid_results = result_df[
            (result_df['Account name'] != 'N/A') & 
            (result_df['Ad group'] != 'No IM_VDP ad groups found') &
            (result_df['Ad group'] != 'No Ad groups with valid Structure is found')
        ]
        
        if len(valid_results) > 0:
            # Get unique account names (Client Name)
            unique_accounts = valid_results['Account name'].unique()
            
            selected_account = st.selectbox(
                "Select an Account to analyze:",
                options=unique_accounts,
                index=0
            )
            
            if selected_account:
                # Filter results for selected account
                account_rows = result_df[result_df['Account name'] == selected_account]
                # Get only rows with valid ad groups
                valid_adgroup_rows = account_rows[
                    (account_rows['Ad group'] != 'No IM_VDP ad groups found') & 
                    (account_rows['Ad group'] != 'No Ad groups with valid Structure is found')
                ]
            
                if len(valid_adgroup_rows) > 0:
                    # Group by ad group name and calculate metrics
                    adgroup_analysis = []
                    
                    # Get unique ad group names for this account
                    unique_adgroups = valid_adgroup_rows['Ad group'].unique()
        
                    for ad_group in unique_adgroups:
                        # Get all rows for this specific ad group
                        adgroup_rows = valid_adgroup_rows[valid_adgroup_rows['Ad group'] == ad_group]
                    
                        # Calculate ads count (how many times this ad group appears)
                        ads_count = len(adgroup_rows)
                        
                        # Get other information from the first occurrence
                        first_row = adgroup_rows.iloc[0]
                        customer_id = first_row['Customer ID']
                        campaign = first_row['Campaign']
                        
                        # Determine ads status - active if any occurrence is active
                        ads_status = 'active' if any(adgroup_rows['ads'] == 'active') else 'not active'
                        
                        # Determine keywords status - active if any occurrence is active
                        keywords_status = 'active' if any(adgroup_rows['keywords'] == 'active') else 'not active'
                        
                        # Calculate keywords count using ad group ID
                        #ad_group_id = None
                        ad_group_ids = []
                        if adgroup_df is not None:
                            # Find the ad group ID using ad group name and customer ID
                            matching_rows = adgroup_df[
                                (adgroup_df['Ad group'] == ad_group) & 
                                (adgroup_df['Customer ID'] == customer_id)
                            ]
                            
                            if not matching_rows.empty:
                                # ad_group_id = matching_rows.iloc[0].get('Ad group ID', None)
                                ad_group_ids = matching_rows['Ad group ID'].tolist()
                            else:
                                ad_group_ids = []

                        # Count keywords for this specific ad group ID
                        # keywords_count = 0
                        # if (
                        #     ad_group_id is not None and 
                        #     pd.notna(ad_group_id) and 
                        #     keyword_df is not None
                        # ):
                        #     keywords_count = len(keyword_df[keyword_df['Ad group ID'] == ad_group_id])
                        keywords_count = 0
                        if ad_group_ids and keyword_df is not None:
                            keywords_count = len(keyword_df[keyword_df['Ad group ID'].isin(ad_group_ids)])
                                                
                        adgroup_analysis.append({
                            'ad_group': ad_group,
                            'campaign': campaign,
                            'ads_status': ads_status,
                            'ads_count': ads_count,
                            'keywords_status': keywords_status,
                            'keywords_count': keywords_count
                        })
                    
                    st.write(f"**Selected Account:** {selected_account}")
                    st.write(f"**Number of Unique Ad Groups:** {len(unique_adgroups)}")
                    
                    # Display analysis for each unique ad group
                    for idx, analysis in enumerate(adgroup_analysis, 1):
                        ad_group = analysis['ad_group']
                        campaign = analysis['campaign']
                        ads_status = analysis['ads_status']
                        ads_count = analysis['ads_count']
                        keywords_status = analysis['keywords_status']
                        keywords_count = analysis['keywords_count']
                        
                        # Display ad group information
                        st.write(f"**Ad Group {idx}:** {ad_group}")
                        if campaign != 'N/A':
                            st.write(f"**Campaign:** {campaign}")
                        
                        # Display metrics in columns
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("Ads Status", ads_status)
                        with col2:
                            st.metric("Ads Count", ads_count)
                        with col3:
                            st.metric("Keywords Status", keywords_status)
                        with col4:
                            st.metric("Keywords Count", keywords_count)
                        
                        st.write("---")  # Separator between ad groups
                    
                    # Account-level summary
                    st.write("### Account Summary:")
                    total_unique_adgroups = len(unique_adgroups)
                    total_ads_count = sum([analysis['ads_count'] for analysis in adgroup_analysis])
                    active_ads_adgroups = len([analysis for analysis in adgroup_analysis if analysis['ads_status'] == 'active'])
                    active_keywords_adgroups = len([analysis for analysis in adgroup_analysis if analysis['keywords_status'] == 'active'])
                    
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Unique Ad Groups", total_unique_adgroups)
                    with col2:
                        st.metric("Total Ads Count", total_ads_count)
                    with col3:
                        st.metric("Ad Groups with Active Ads", f"{active_ads_adgroups}/{total_unique_adgroups}")
                    with col4:
                        st.metric("Ad Groups with Active Keywords", f"{active_keywords_adgroups}/{total_unique_adgroups}")
                        
                else:
                    st.info(f"No valid ad groups found for account: {selected_account}")
        else:
            st.info("No valid accounts found for analysis.")
        
       
    else:
        st.info("No records found matching the criteria.")